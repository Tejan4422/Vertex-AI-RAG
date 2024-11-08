import functions_framework
from google.cloud import storage
import pandas as pd
import json
import time
from vertexai.generative_models import GenerativeModel, ChatSession
from google.cloud import discoveryengine_v1 as discoveryengine
import vertexai
from io import BytesIO
import requests
from google.auth import default
from google.auth.transport.requests import Request
from typing import List
from google.api_core.client_options import ClientOptions
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from google.api_core.exceptions import InternalServerError
import logging

# Initialize Vertex AI and Gemini model
def initialize_vertex_ai():
    project_id = ""  # Update with your project ID
    vertexai.init(project=project_id, location="")
    model = GenerativeModel("gemini-1.0-pro")
    return model.start_chat()

# Initialize Vertex AI and start chat session


chat = initialize_vertex_ai()
storage_client = storage.Client(project="")  

# Get the access token for authentication
def get_access_token():
    credentials, _ = default()
    credentials.refresh(Request())
    return credentials.token

# Function to get chat response from the Gemini model
#def get_chat_response(chat: ChatSession, prompt: str) -> str:
#    text_response = []
#    responses = chat.send_message(prompt, stream=True)
#    for chunk in responses:
#        text_response.append(chunk.text)
#    return "".join(text_response)


def get_chat_response(chat: ChatSession, prompt: str) -> str:

      text_response = []
      retry_count = 0
      max_retries = 3  # Adjust as needed
      while retry_count < max_retries:
          try:
              responses = chat.send_message(prompt, stream=True)
              for chunk in responses:
                  text_response.append(chunk.text)
              return "".join(text_response)
          except InternalServerError as e:
              retry_count += 1
              wait_time = 2 ** retry_count  # Exponential backoff
              print(f"InternalServerError encountered. Retrying in {wait_time} seconds...")
              time.sleep(wait_time)
      print(f"Max retries reached. Failed to get a response for prompt: {prompt}")
      return "" # or raise an exception


"""
# Trying Logging and enhance cloud function
logging.basicConfig(level=logging.INFO)  # Configure basic logging
def get_chat_response(chat: ChatSession, prompt: str) -> str:
    text_response = []
    retry_count = 0
    max_retries = 3  # Adjust as needed
    while retry_count < max_retries:
        try:
            responses = chat.send_message(prompt, stream=True)
            for chunk in responses:
                text_response.append(chunk.text)
            return "".join(text_response)
        except (InternalServerError, ResourceExhausted, ServiceUnavailable) as e:
            retry_count += 1
            wait_time = 2 ** retry_count  # Exponential backoff
            logging.warning(f"Error encountered: {str(e)}. Retrying in {wait_time} seconds...")
            time.sleep(wait_time)
        except Exception as e:
            logging.error(f"Unexpected error: {str(e)}")
            break  # Break on unknown errors after logging
    
    logging.error(f"Max retries reached. Failed to get a response for prompt: {prompt}")
    return ""  # or raise an exception
"""


# Function to call the Discovery Engine API
def call_discovery_engine(query_text, query_id=""):
    url = "https://discoveryengine.googleapis.com/v1alpha/projects/994544669142/locations/global/collections/default_collection/engines/rfp-agent_1722488765661/servingConfigs/default_search:answer"

    access_token = get_access_token()
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }

    payload = {
        "query": {
            "text": query_text,
            "queryId": query_id
        },
        "session": "",
        "answerGenerationSpec": {
            "ignoreAdversarialQuery": True,
            "ignoreNonAnswerSeekingQuery": True,
            "ignoreLowRelevantContent": True,
            "includeCitations": True,
            "promptSpec": {
                "preamble": (
                    "Please keep the answer concise and limit it to between 100 to 200 words."

                ),
            },
            "modelSpec": {
                "modelVersion": "gemini-1.0-pro-002/answer_gen/v1"
            }
        }
    }
    response = requests.post(url, headers=headers, data=json.dumps(payload))

    if response.status_code == 200:
        return response.json()
    else:
        raise Exception(f"Request failed with status code {response.status_code}: {response.text}")

"""
def call_discovery_engine(query_text, query_id=""):
    url = "https://discoveryengine.googleapis.com/v1alpha/projects/994544669142/locations/global/collections/default_collection/engines/rfp-agent_1722488765661/servingConfigs/default_search:answer"
    access_token = get_access_token()
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }

    payload = {
        "query": {
            "text": query_text,
            "queryId": query_id
        },
        "session": "",
        "answerGenerationSpec": {
            "ignoreAdversarialQuery": True,
            "ignoreNonAnswerSeekingQuery": True,
            "ignoreLowRelevantContent": True,
            "includeCitations": True,
            "promptSpec": {
                "preamble": (
                    "Please keep the answer concise and limit it to between 100 to 200 words."
                ),
            },
            "modelSpec": {
                "modelVersion": "gemini-1.0-pro-002/answer_gen/v1"
            }
        }
    }
    
    retry_count = 0
    max_retries = 3
    while retry_count < max_retries:
        try:
            response = requests.post(url, headers=headers, data=json.dumps(payload))
            if response.status_code == 200:
                return response.json()
            else:
                raise Exception(f"Request failed with status code {response.status_code}: {response.text}")
        except requests.exceptions.RequestException as e:
            retry_count += 1
            wait_time = 2 ** retry_count  # Exponential backoff
            print(f"Error: {str(e)}. Retrying in {wait_time} seconds...")
            time.sleep(wait_time)
    
    print("Max retries reached. Failed to get a response from Discovery Engine.")
    return {}  # Return an empty dictionary or handle it appropriately in the calling function

"""

# Function to extract relevant information from the Discovery Engine API response
def extract_relevant_info(response_json):
    answer_text = response_json.get("answer", {}).get("answerText", "")
    uris = []
    titles = []

    references = response_json.get("answer", {}).get("references", [])
    for reference in references:
        chunk_info = reference.get("chunkInfo", {})
        document_metadata = chunk_info.get("documentMetadata", {})
        uris.append(document_metadata.get("uri", ""))
        titles.append(document_metadata.get("title", ""))

    return {
        "answerText": answer_text,
        "uris": uris,
        "titles": titles[:3]  # Only taking top 3 titles
    }



# Function to format response
def format_response(response_text: str) -> str:
    formatted_text = response_text.replace('*', '').replace('**', '')
    formatted_text = formatted_text.replace('**', '').replace('**', '')
    formatted_text = formatted_text.replace('***', '').replace('***', '')

    return formatted_text

# Processing each sheet with conditions and handling blank values
def process_each_sheet_v2(df: pd.DataFrame, chat: ChatSession, requirements_col: str) -> pd.DataFrame:
    # Function to check if a row should be processed or ignored (e.g., section names or empty rows)
    def should_process(row):
        value = row[requirements_col]
        if pd.isna(value) or len(str(value).split()) <= 3:  # Ignore short headings or empty rows
            return False
        return True

    # Function to process the questions and RAG responses
    def process_question(question):
        try:
            response_json = call_discovery_engine(question)
            relevant_info = extract_relevant_info(response_json)
            answer_text = relevant_info.get("answerText", "")
            titles = relevant_info.get("titles", [])[:3]
            formatted_answer_text = format_response(answer_text)
            references = ", ".join(titles)
            time.sleep(20)
            return formatted_answer_text, references
        except Exception as e:
            print(f"Error processing question '{question}': {str(e)}")
            return "", ""

    # Apply the logic to transform only the rows that should be processed
    df['Questions'] = df.apply(lambda row: get_chat_response(chat, f"Please transform the following requirement into a clear Yes/No question. The question should directly check whether the system supports the requirement. Requirement: {row[requirements_col]}") 
                                    if should_process(row) else None, axis=1)
    
    # Apply RAG processing only to rows that have a valid question
    df[['Vendor Response', 'References']] = df.apply(lambda row: process_question(row['Questions']) 
                                                     if pd.notna(row['Questions']) else ("", ""), axis=1).apply(pd.Series)

    return df

# Apply formatting to sheet
def apply_formatting(sheet):
    header_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

# Cloud Function entry point
@functions_framework.cloud_event
def hello_gcs(cloud_event):
    try:
        # Extract relevant information from the event
        data = cloud_event.data
        bucket_name = data["bucket"]
        file_name = data["name"]

        # Skip processing if the file has already been processed
        if file_name.startswith("processed_"):
            print(f"File '{file_name}' has already been processed. Skipping...")
            return

        # Initialize Cloud Storage client
        storage_client = storage.Client()
        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(file_name)

        # Download the Excel file from the bucket
        downloaded_blob = blob.download_as_bytes()
        xls = pd.ExcelFile(BytesIO(downloaded_blob))

        # Process each sheet and store the results
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)

                # Check the structure of the sheet to determine the correct column to process
                if 'Bank Requirement' in df.columns:
                    processed_df = process_each_sheet_v2(df, chat, 'Bank Requirement')
                elif 'requirements' in df.columns:
                    processed_df = process_each_sheet_v2(df, chat, 'requirements')
                else:
                    continue  # Skip sheets that do not have the expected columns

                processed_df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Access the worksheet and apply formatting
                workbook = writer.book
                worksheet = workbook[sheet_name]
                apply_formatting(worksheet)

        output.seek(0)  # Reset the stream position before uploading

        # Upload the processed file to GCS
        output_filename = f"processed_{file_name}"
        output_blob = bucket.blob(output_filename)
        output_blob.upload_from_file(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        raise
